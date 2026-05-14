import json
import openpyxl
from pathlib import Path
import os

def flatten_dict(d, parent_key='', sep='.'):
    """Convert nested dict to flat dict with dot-separated keys"""
    items = []
    for k, v in d.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, dict):
            items.extend(flatten_dict(v, new_key, sep=sep).items())
        else:
            items.append((new_key, v))
    return dict(items)

# Load JSON files
script_dir = Path(__file__).parent
project_root = script_dir.parent

with open(project_root / 'frontend/src/locales/en.json', 'r', encoding='utf-8') as f:
    en_data = json.load(f)

with open(project_root / 'frontend/src/locales/vi.json', 'r', encoding='utf-8') as f:
    vi_data = json.load(f)

# Flatten dictionaries
en_flat = flatten_dict(en_data)
vi_flat = flatten_dict(vi_data)

# Load Excel workbook
wb = openpyxl.load_workbook(script_dir / 'languages.xlsx')
ws = wb.active

# Get existing keys from Excel
existing_keys = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    if row and row[0]:
        existing_keys.add(str(row[0]))

# Find the last row
last_row = ws.max_row

# Add new keys to Excel
for key in sorted(en_flat.keys()):
    if key not in existing_keys:
        last_row += 1
        ws.cell(row=last_row, column=1).value = key
        ws.cell(row=last_row, column=2).value = en_flat.get(key, '')
        ws.cell(row=last_row, column=3).value = vi_flat.get(key, '')
        print(f"Added: {key}")

# Update existing keys
for row in ws.iter_rows(min_row=2):
    key = row[0].value
    if key and key in en_flat:
        row[1].value = en_flat[key]
        row[2].value = vi_flat.get(key, '')

# Save Excel file
wb.save(script_dir / 'languages.xlsx')
print("✅ Excel file updated successfully")
